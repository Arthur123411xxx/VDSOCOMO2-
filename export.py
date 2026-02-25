"""
export.py - Export des résultats en Excel multi-onglets et CSV
"""

import json
import pandas as pd
from io import BytesIO
from typing import List
from datetime import datetime
from zipfile import ZipFile, ZIP_DEFLATED

import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

import csv


def _to_float(value):
    """Convertit '7 894,66' / '1.248,000' / 995.904 / 995,904 en float si possible."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    s = s.replace("\u00a0", " ").replace(" ", "")
    # si on a virgule et point, on suppose '.' milliers et ',' décimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except:
            return None

    # si on a uniquement des virgules -> décimal
    if "," in s:
        try:
            return float(s.replace(",", "."))
        except:
            return None

    # uniquement points: soit milliers, soit décimal — on tente direct
    try:
        return float(s)
    except:
        try:
            return float(s.replace(".", ""))
        except:
            return None


def _format_ws(ws, money_cols=None, percent_cols=None, wrap_cols=None):
    """Mise en forme de base : en-tête, filtres, gel, largeur auto, formats."""
    money_cols = money_cols or []
    percent_cols = percent_cols or []
    wrap_cols = wrap_cols or []

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    from openpyxl.styles import Font

    def _apply_pdf_hyperlinks(ws, base_rel_prefix="../", pe_header="PE", file_header="Fichier", path_header="PDF path"):
        # map headers -> column index
        headers = {c.value: idx + 1 for idx, c in enumerate(ws[1])}
        col_pe = headers.get(pe_header)
        col_file = headers.get(file_header)
        col_path = headers.get(path_header)
        if not col_path:
            return

        # hide the path column (report stays clean)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_path)].hidden = True

        for r in range(2, ws.max_row + 1):
            path = ws.cell(row=r, column=col_path).value
            if not path:
                continue

            rel = (base_rel_prefix + str(path)).replace("\\", "/")

            # priority: hyperlink on PE; fallback on filename if PE empty
            target_col = col_pe if (col_pe and ws.cell(r, col_pe).value) else col_file
            if not target_col:
                continue

            cell = ws.cell(row=r, column=target_col)
            cell.hyperlink = rel
            cell.style = "Hyperlink"

    # header style
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # largeur auto
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 55)

    # formats
    headers = [c.value for c in ws[1]]
    for j, h in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        if h in money_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = '#,##0.00 "€"'
        if h in percent_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = '0.00"%"'
        if h in wrap_cols:
            for cell in ws[col_letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def export_to_excel(invoices_data: List[dict],
                   referential_df: pd.DataFrame = None) -> BytesIO:
    """
    Excel minimal et pro.

    Onglets:
    - COMPARAISON : lignes OCR vs CSV (inv["line_comparison"])
      + Δ Total (€)
      + Source/Note (ex: MANQUANT_OCR => "CSV_ONLY")
    - PE_VALIDE : PE où toutes les lignes ont Statut ligne = OK
    - PE_PROBLEME : PE où au moins une ligne n'est pas OK (INFO/ECART/MANQUANT_OCR/etc.)
    """
    output = BytesIO()

    # ----------------------------
    # 1) COMPARAISON (lignes)
    # ----------------------------
    lines_rows = []

    for inv in invoices_data:
        pe = inv.get("pe_selected") or inv.get("pe") or ""
        ac = inv.get("ac", "")

        flags_fact = inv.get("flags") or []
        ac_raw = (inv.get("ac_raw") or "").upper()
        ac_type = (inv.get("ac_type") or "").upper()

        # ✅ True si l'OCR a lu AM (ou si le flag a été ajouté à l'extraction)
        flag_am = (ac_type == "AM") or ac_raw.startswith("AM") or ("AC_LU_COMME_AM" in flags_fact)

        for r in (inv.get("line_comparison") or []):
            qty_ocr = _to_float(r.get("qty_ocr"))
            qty_ref = _to_float(r.get("qty_ref"))
            pu_ocr  = _to_float(r.get("pu_ocr"))
            pu_ref  = _to_float(r.get("pu_ref"))
            tot_ocr = _to_float(r.get("total_ocr"))
            tot_ref = _to_float(r.get("total_ref"))

            delta_qty   = (qty_ocr - qty_ref) if (qty_ocr is not None and qty_ref is not None) else None
            delta_pu    = (pu_ocr - pu_ref)   if (pu_ocr  is not None and pu_ref  is not None) else None
            delta_total = (tot_ocr - tot_ref) if (tot_ocr is not None and tot_ref is not None) else None

            st_line = str(r.get("status") or "").strip().upper()
            source = "CSV_ONLY" if st_line == "MANQUANT_OCR" else "MATCH"
            note = "Présent en CSV, absent OCR" if st_line == "MANQUANT_OCR" else ""

            lines_rows.append({
                "PE": pe,
                "PDF path": inv.get("zip_pdf_path", ""),
                "AC": ac,
                "AC brut": inv.get("ac_raw", ""),
                "Type AC": inv.get("ac_type", ""),
                "Flag AM": "OUI" if flag_am else "",
                "code_article": r.get("code_article"),
                "desc_ocr": r.get("desc_ocr") or "",
                "Source": source,
                "Note": note,

                "|": "",

                "Qté OCR": qty_ocr,
                "PU OCR": pu_ocr,
                "Total OCR": tot_ocr,

                "||": "",

                "Qté CSV": qty_ref,
                "PU CSV": pu_ref,
                "Total CSV": tot_ref,

                "|||": "",

                "Δ Qté (u)": delta_qty,
                "Δ PU (€)": delta_pu,
                "Δ Total (€)": delta_total,

                "||||": "",

                "Statut Qté": r.get("status_qty") or "",
                "Statut PU": r.get("status_pu") or "",
                "Statut total": r.get("status_total") or "",
                "Statut ligne": r.get("status") or "",
            })

    df_lines = pd.DataFrame(lines_rows)

    ordered_cols = [
        "PE","PDF path","AC", "AC brut", "Type AC", "Flag AM", "code_article", "desc_ocr", "Source", "Note",
        "|",
        "Qté OCR", "PU OCR", "Total OCR",
        "||",
        "Qté CSV", "PU CSV", "Total CSV",
        "|||",
        "Δ Qté (u)", "Δ PU (€)", "Δ Total (€)",
        "||||",
        "Statut Qté", "Statut PU", "Statut total", "Statut ligne",
    ]

    if not df_lines.empty:
        df_lines = df_lines.reindex(columns=[c for c in ordered_cols if c in df_lines.columns])

        # Tri par PE numérique puis code_article
        df_lines["PE_num"] = (
            df_lines["PE"].astype(str)
            .str.replace(r"\D+", "", regex=True)
            .replace("", "0")
            .astype(int)
        )
        df_lines = (
            df_lines.sort_values(by=["PE_num", "PE", "code_article"], kind="mergesort")
            .drop(columns=["PE_num"])
        )

    # ----------------------------
    # 2) PE_VALIDE / PE_PROBLEME (basé UNIQUEMENT sur Statut ligne)
    # ----------------------------
    pe_rows = []
    pe_is_valid = {}

    for inv in invoices_data:
        pe = inv.get("pe_selected") or inv.get("pe") or ""
        if not pe:
            continue

        line_cmp = inv.get("line_comparison") or []

        if not line_cmp:
            pe_is_valid[pe] = False
            status_summary = "AUCUNE_LIGNE"
        else:
            statuses = [str(x.get("status") or "").strip().upper() for x in line_cmp]
            is_valid = all(st == "OK" for st in statuses)
            pe_is_valid[pe] = pe_is_valid.get(pe, True) and is_valid
            uniq = sorted({st for st in statuses if st})
            status_summary = ", ".join(uniq) if uniq else ""

        pe_rows.append({
            "PE": pe,
            "PDF path": inv.get("zip_pdf_path", ""),
            "AC": inv.get("ac", ""),
            "Fichier": inv.get("source_filename") or inv.get("filename") or "",
            "Statuts lignes (uniq)": status_summary,
            "Nb lignes": len(line_cmp),
            "Total facture": _to_float(inv.get("total_facture") or inv.get("invoice_total") or inv.get("total_amount")),
            "Attendu (CSV)": _to_float(inv.get("montant_total_attendu") or inv.get("expected_total")),
            "Delta (€)": _to_float(inv.get("delta")),
            "Delta (%)": _to_float(inv.get("delta_pct")),
        })

    df_pe = pd.DataFrame(pe_rows)

    if not df_pe.empty:
        df_pe = df_pe.drop_duplicates(subset=["PE"], keep="last")
        df_pe["__is_valid"] = df_pe["PE"].map(lambda x: bool(pe_is_valid.get(x, False)))

        df_ok = df_pe[df_pe["__is_valid"]].drop(columns=["__is_valid"])
        df_bad = df_pe[~df_pe["__is_valid"]].drop(columns=["__is_valid"])

        def _pe_num_series(s):
            return s.astype(str).str.replace(r"\D+", "", regex=True).replace("", "0").astype(int)

        df_ok = df_ok.assign(PE_num=_pe_num_series(df_ok["PE"])).sort_values(["PE_num", "PE"], kind="mergesort").drop(columns=["PE_num"])
        df_bad = df_bad.assign(PE_num=_pe_num_series(df_bad["PE"])).sort_values(["PE_num", "PE"], kind="mergesort").drop(columns=["PE_num"])
    else:
        df_ok = pd.DataFrame(columns=["PE"])
        df_bad = pd.DataFrame(columns=["PE"])

    # ----------------------------
    # 3) Écriture Excel
    # ----------------------------
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_lines.to_excel(writer, sheet_name="COMPARAISON", index=False)
        df_ok.to_excel(writer, sheet_name="PE_VALIDE", index=False)
        df_bad.to_excel(writer, sheet_name="PE_PROBLEME", index=False)

        for name in ["COMPARAISON", "PE_VALIDE", "PE_PROBLEME"]:
            ws = writer.sheets[name]
            _format_ws(
                ws,
                money_cols=[
                    "PU OCR", "Total OCR", "PU CSV", "Total CSV",
                    "Δ PU (€)", "Δ Total (€)",
                    "Total facture", "Attendu (CSV)", "Delta (€)",
                ],
                percent_cols=["Delta (%)"],
                wrap_cols=["desc_ocr", "Statuts lignes (uniq)", "Note"]
            )
            _apply_pdf_hyperlinks(ws)

    output.seek(0)
    return output

from openpyxl.styles import Font

import os
from openpyxl.utils import get_column_letter

def _apply_pdf_hyperlinks(ws, pe_header="PE", file_header="Fichier", path_header="PDF path"):
    headers = {c.value: idx + 1 for idx, c in enumerate(ws[1])}
    col_pe = headers.get(pe_header)
    col_file = headers.get(file_header)
    col_path = headers.get(path_header)
    if not col_path:
        return

    # cache la colonne technique
    ws.column_dimensions[get_column_letter(col_path)].hidden = True

    base = ".." + os.sep  # "../" sur Mac, "..\" sur Windows

    for r in range(2, ws.max_row + 1):
        path = ws.cell(row=r, column=col_path).value
        if not path:
            continue

        p = str(path).lstrip("/\\")
        p = p.replace("/", os.sep).replace("\\", os.sep)

        rel = base + p

        target_col = col_pe if (col_pe and ws.cell(r, col_pe).value) else col_file
        if not target_col:
            continue

        cell = ws.cell(row=r, column=target_col)
        cell.hyperlink = rel
        cell.style = "Hyperlink"




def export_to_csv(invoices_data: List[dict], csv_folder_in_zip: str = "CSV") -> BytesIO:
    """
    Génère un fichier CSV des résultats + une colonne pdf_path
    pdf_path pointe vers le PDF dans le ZIP (chemin relatif).

    csv_folder_in_zip = "CSV" signifie que le CSV sera placé dans le dossier CSV/ du ZIP
    => le lien doit remonter d'un niveau : ../OK/PE123456.pdf
    """
    rows = []

    # si le CSV est dans un sous-dossier, il faut remonter vers la racine du zip
    prefix = "../" if csv_folder_in_zip else ""

    for inv in invoices_data:
        statut = (inv.get("statut_final") or "UNKNOWN").upper()

        pe = inv.get("pe_selected") or inv.get("pe") or ""
        source = (inv.get("source_filename") or inv.get("filename") or "facture.pdf")
        source = source.replace("\\", "_").replace("/", "_")

        # IMPORTANT: on doit suivre exactement la même règle de nommage que tes PDFs dans le ZIP
        if statut in ["OK", "ECART"] and pe:
            pdf_name = f"{pe}.pdf"
        else:
            pe_safe = pe if pe else "PE_INCONNU"
            pdf_name = f"{pe_safe}__{source}"
            if not pdf_name.lower().endswith(".pdf"):
                pdf_name += ".pdf"

        pdf_path = f"{prefix}{statut}/{pdf_name}" if statut else ""

        row = {
            'fichier': inv.get('filename', ''),
            'pe': inv.get('pe_selected', ''),
            'pe_status': inv.get('pe_status', ''),
            'total_facture': inv.get('total_facture'),
            'montant_attendu': inv.get('montant_total_attendu'),
            'delta': inv.get('delta'),
            'delta_pct': inv.get('delta_pct'),
            'statut': inv.get('statut_final', ''),
            'fournisseur': inv.get('supplier', ''),
            'numero_facture': inv.get('invoice_number', ''),
            'date_facture': inv.get('invoice_date', ''),
            'devise': inv.get('currency', 'EUR'),
            'flags': '|'.join(inv.get('flags', [])),
            'warnings': '|'.join(inv.get('warnings', [])),
            'pdf_path': pdf_path
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    output = BytesIO()

    # --- Nettoyage / arrondis pour lisibilité ---
    for col in ["unites_detail", "prix_unitaire", "total_ligne", "total_facture"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(3)

    # --- Contrôle de cohérence ligne ---
    def calc_diff(row):
        try:
            if pd.notna(row.get("unites_detail")) and pd.notna(row.get("prix_unitaire")) and pd.notna(row.get("total_ligne")):
                return (row["unites_detail"] * row["prix_unitaire"]) - row["total_ligne"]
            if pd.notna(row.get("quantite")) and pd.notna(row.get("prix_unitaire")) and pd.notna(row.get("total_ligne")):
                return (row["quantite"] * row["prix_unitaire"]) - row["total_ligne"]
        except Exception:
            return None
        return None

    df["diff_calc_total"] = df.apply(calc_diff, axis=1)
    if "diff_calc_total" in df.columns:
        df["diff_calc_total"] = pd.to_numeric(df["diff_calc_total"], errors="coerce").round(3)

    sort_cols = [c for c in ["pe", "fichier", "ligne_index"] if c in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols)

    df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    return output


def export_items_to_csv(invoices_data: List[dict], csv_folder_in_zip: str = "CSV") -> BytesIO:
    """
    1 ligne = 1 ligne article extraite (items).
    """
    rows = []

    for inv in invoices_data:
        items_block = inv.get("items") or {}
        items = items_block.get("items") or []

        for idx, it in enumerate(items, start=1):
            rows.append({
                "fichier": inv.get("filename", ""),
                "pe": inv.get("pe_selected", ""),
                "statut": inv.get("statut_final", ""),
                "total_facture": inv.get("total_facture"),

                "ligne_index": idx,
                "code": it.get("item_code"),
                "description": it.get("description"),
                "quantite": it.get("quantity"),
                "unite": it.get("unit"),
                "unites_detail": it.get("units_detail"),
                "uom_detail": it.get("units_detail_uom"),
                "prix_unitaire": it.get("unit_price"),
                "total_ligne": it.get("line_total"),
                "ok": it.get("ok"),
                "raw": it.get("raw_line") or it.get("raw_item_line"),
            })

    df = pd.DataFrame(rows)
    output = BytesIO()

    df.to_csv(
        output,
        index=False,
        sep=";",  # ✅ séparateur Excel BE/FR
        decimal=",",  # ✅ décimales FR
        encoding="utf-8-sig",
        quoting=csv.QUOTE_ALL  # ✅ protège les virgules dans le texte
    )
    output.seek(0)
    return output


def generate_report_filename(prefix: str = "export_VD", date_only: bool = True) -> str:
    """
    Génère un nom de fichier avec date (et optionnellement l'heure).
    """
    if date_only:
        stamp = datetime.now().strftime("%Y-%m-%d")   # ex: 2026-02-11
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}"


from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO
import re


def export_pdfs_zip_by_status(invoices_data) -> BytesIO:
    """
    ZIP unique avec sous-dossiers:
    - ECART/
    - PE_INCONNU/
    - MANQUANT_PE/
    (et autres statuts si présents)
    """
    out = BytesIO()

    def safe(s: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]", "", str(s))[:60] or "UNKNOWN"

    with ZipFile(out, "w", compression=ZIP_DEFLATED) as zf:
        existing = set()

        for inv in invoices_data:
            statut = (inv.get("statut_final") or "UNKNOWN").upper()
            pdf_bytes = inv.get("pdf_bytes")
            if not pdf_bytes:
                continue

            pe = inv.get("pe_selected") or "PE_INCONNU"
            pe_clean = safe(pe)

            original = inv.get("source_filename") or inv.get("filename") or "facture.pdf"
            original = original.replace("\\", "_").replace("/", "_")

            # Nom selon statut
            if statut == "ECART":
                base_name = f"{pe_clean}.pdf"
            else:
                base_name = f"{pe_clean}__{original}"
                if not base_name.lower().endswith(".pdf"):
                    base_name += ".pdf"

            zip_name = f"{statut}/{base_name}"

            if zip_name in existing:
                i = 2
                while f"{statut}/{base_name[:-4]}_{i}.pdf" in existing:
                    i += 1
                zip_name = f"{statut}/{base_name[:-4]}_{i}.pdf"

            existing.add(zip_name)
            zf.writestr(zip_name, pdf_bytes)

    out.seek(0)
    return out


from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO
import pandas as pd


def make_ok_invoices_excel(invoices_data) -> BytesIO:
    """
    Excel 'PE validées (OK)' :
    PE + prix facture + prix vandamme + écart accepté + % + fichier source.
    """
    # NOTE: fonction laissée telle quelle (elle était vide chez toi).
    return None


def make_summary_excel(invoices, title="RESUME") -> BytesIO:
    rows = []
    for inv in invoices:
        rows.append({
            "PE": inv.get("pe_selected") or inv.get("pe") or "",
            "Fichier": inv.get("source_filename") or inv.get("filename") or "",
            "Statut": inv.get("statut_final") or "",
            "Flags": ", ".join(inv.get("flags") or []),

            "Total facture": inv.get("total_facture") or inv.get("invoice_total") or inv.get("total_amount"),
            "Attendu (CSV)": inv.get("montant_total_attendu") or inv.get("expected_total"),
            "Delta (€)": inv.get("delta"),
            "Delta (%)": inv.get("delta_pct"),

            "Qté OCR": inv.get("quantite_total_ocr"),
            "Unité OCR": inv.get("quantite_unite_ocr") or "",
            "Qté (CSV)": inv.get("quantite_totale_attendue"),
            "Delta Qté": inv.get("delta_qte"),
            "Statut Qté": inv.get("quantite_status") or "",
        })

    df = pd.DataFrame(rows)
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=title[:31])
    out.seek(0)
    return out


def make_price_error_excel_for_invoice(inv: dict) -> BytesIO:
    items_block = inv.get("items") or {}
    items = items_block.get("items") or []

    rows = []
    for idx, it in enumerate(items, start=1):
        if (it.get("item_price_status") or "").upper() != "ECART":
            continue

        rows.append({
            "PE": inv.get("pe_selected") or inv.get("pe") or "",
            "Fichier": inv.get("source_filename") or inv.get("filename") or "",
            "Date facture": inv.get("invoice_date") or "",
            "Flags facture": ", ".join(inv.get("flags", []) or []),

            "Ligne": idx,
            "Code SOCOMO": it.get("item_code"),
            "Description": it.get("description"),
            "Qté": it.get("quantity"),
            "Unité": it.get("unit"),
            "PU (SOCOMO)": it.get("unit_price"),
            "PU attendu (VANDAMME)": it.get("expected_unit_price"),
            "Δ PU (€)": it.get("delta_unit"),
            "Δ PU %": it.get("delta_unit_pct"),
            "Total ligne": it.get("line_total"),
        })

    df = pd.DataFrame(rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ECARTS_PRIX", index=False)
    output.seek(0)
    return output


def make_total_qty_error_excel_for_invoice(inv: dict) -> BytesIO:
    pe = inv.get("pe_selected") or inv.get("pe") or ""

    rows = [{
        "PE": pe,
        "Fichier": inv.get("source_filename") or inv.get("filename") or "",
        "Date facture": inv.get("invoice_date") or "",
        "Statut facture": inv.get("statut_final") or "",
        "Flags": ", ".join(inv.get("flags") or []),

        "Total facture": _to_float(inv.get("total_facture") or inv.get("invoice_total") or inv.get("total_amount")),
        "Attendu (CSV)": _to_float(inv.get("montant_total_attendu") or inv.get("expected_total")),
        "Delta (€)": _to_float(inv.get("delta")),
        "Delta (%)": _to_float(inv.get("delta_pct")),

        "Qté OCR": _to_float(inv.get("quantite_total_ocr")),
        "Unité OCR": inv.get("quantite_unite_ocr") or "",
        "Qté (CSV)": _to_float(inv.get("quantite_totale_attendue")),
        "Delta Qté": _to_float(inv.get("delta_qte")),
        "Statut Qté": inv.get("quantite_status") or "",
        "Détail Qté OCR": json.dumps(inv.get("quantite_by_unit_ocr") or {}, ensure_ascii=False),
    }]

    df = pd.DataFrame(rows)
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ERREUR", index=False)
    out.seek(0)
    return out


def export_bundle_zip(
    invoices_data,
    include_full_excel=True,
    include_extraits_excel=True,
    include_pdfs=True,
    include_erreurs_par_pe=True,
    include_csv=True,
) -> BytesIO:
    """
    ZIP propre et complet:
    01_RAPPORT/rapport_complet.xlsx + extraits + README
    02_PDF/<STATUT>/PE.pdf
    03_ERREURS_PAR_PE/PE.../ (pdf + excels d'erreurs)
    04_CSV/ (lignes_articles + erreurs)
    """
    out = BytesIO()

    def safe(s: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]", "", str(s))[:80] or "UNKNOWN"

    def pe_of(inv: dict) -> str:
        return inv.get("pe_selected") or inv.get("pe") or ""

    def pdf_suffix_am(inv: dict) -> str:
        flags = inv.get("flags") or []
        ac_type = (inv.get("ac_type") or "").upper()
        ac_raw = (inv.get("ac_raw") or inv.get("ac") or "").upper()

        is_am = (
                ac_type == "AM"
                or ac_raw.startswith("AM")
                or ("AC_LU_COMME_AM" in flags)
                or ("AM_AU_LIEU_DE_AC" in flags)
        )
        if not is_am:
            return ""

        digits = re.sub(r"\D+", "", ac_raw)
        return f"__AM{digits}" if digits else "__AM"

    # ✅ Pré-calcule le nom exact des PDFs dans le ZIP (gère _2, _3...)
    if include_pdfs:
        existing = set()
        idx_no_pe = 1

        for inv in invoices_data:
            pdf_bytes = inv.get("pdf_bytes")
            if not pdf_bytes:
                inv["zip_pdf_path"] = ""
                continue

            pe = pe_of(inv)

            if pe:
                zip_name = f"02_PDF/{safe(pe)}{pdf_suffix_am(inv)}.pdf"
            else:
                zip_name = f"02_PDF/SANS_PE/facture_{idx_no_pe:03d}.pdf"
                idx_no_pe += 1

            if zip_name in existing:
                base = zip_name[:-4]
                i = 2
                while f"{base}_{i}.pdf" in existing:
                    i += 1
                zip_name = f"{base}_{i}.pdf"
            existing.add(zip_name)

            inv["zip_pdf_path"] = zip_name  # 👈 on stocke le chemin exact

    def classify_pdf_folder(inv: dict) -> str:
        st = (inv.get("statut_final") or "").upper()
        if not pe_of(inv):
            return "SANS_PE"
        if st == "OK":
            return "OK"
        if st == "ECART_LIGNES":
            return "ECART_LIGNES"
        if st == "ECART":
            return "ECART_TOTAL"
        return "AUTRES"

    def has_price_issue(inv: dict) -> bool:
        flags = inv.get("flags", []) or []
        return ("PRIX_ARTICLE_ECART" in flags) or ((inv.get("items_ecarts") or 0) > 0)

    def has_qty_issue(inv: dict) -> bool:
        qs = (inv.get("quantite_status") or "").upper()
        flags = inv.get("flags", []) or []
        return ("ECART_QTE" in flags) or qs in ("ECART_QTE", "MANQUANT_QTE_OCR", "MANQUANT_QTE_REF", "ERREUR_QTE")

    def has_total_issue(inv: dict) -> bool:
        st = (inv.get("statut_final") or "").upper()
        d = _to_float(inv.get("delta"))
        if st == "ECART":
            return True
        if d is None:
            return False
        return abs(d) > 0.01

    def build_errors_rows():
        rows = []
        for inv in invoices_data:
            pe = pe_of(inv)
            if not pe:
                continue

            reasons = set()
            line_cmp = inv.get("line_comparison") or []
            for r in line_cmp:
                if (r.get("status_qty") or "") != "OK":
                    reasons.add("QUANTITE")
                if (r.get("status_pu") or "") != "OK":
                    reasons.add("PRIX")
                if (r.get("status_total") or "") != "OK":
                    reasons.add("TOTAL")

            if (inv.get("quantite_status") or "") == "ECART":
                reasons.add("QUANTITE")

            if (inv.get("statut_final") or "") in ["ECART", "ECART_LIGNES"]:
                reasons.add("TOTAL")

            flags = inv.get("flags") or []
            if not reasons and ((inv.get("statut_final") or "") != "OK" or flags):
                reasons.add("AUTRE")

            for reason in sorted(reasons):
                rows.append({
                    "PE": pe,
                    "Erreur": reason,
                    "Fichier": inv.get("filename", ""),
                    "Statut final": inv.get("statut_final", ""),
                    "Nb lignes comparees": len(line_cmp),
                    "Flags": ", ".join(flags),
                })
        return rows

    with ZipFile(out, "w", compression=ZIP_DEFLATED) as zf:
        readme = (
            "README - Contenu du ZIP\n"
            "01_RAPPORT/rapport_complet.xlsx : rapport complet (FACTURES + LIGNES + ERREURS ...)\n"
            "02_PDF/ : PDFs classés par statut\n"
        )
        zf.writestr("01_RAPPORT/README.txt", readme)

        if include_full_excel:
            full_buf = export_to_excel(invoices_data)
            zf.writestr("01_RAPPORT/rapport_complet.xlsx", full_buf.getvalue())

        if include_extraits_excel:
            ok_buf = make_ok_invoices_excel(invoices_data)

            if ok_buf is None:
                pass
            else:
                ok_bytes = ok_buf.getvalue() if hasattr(ok_buf, "getvalue") else ok_buf
                zf.writestr("01_RAPPORT/extraits/PE_valides.xlsx", ok_bytes)

            problems_qty_price = [inv for inv in invoices_data if has_qty_issue(inv) or has_price_issue(inv)]
            problems_total = [inv for inv in invoices_data if has_total_issue(inv)]

            if problems_qty_price:
                buf = make_summary_excel(problems_qty_price, title="QTE_PRIX")
                zf.writestr("01_RAPPORT/extraits/PE_problemes_qte_prix.xlsx", buf.getvalue())

            if problems_total:
                buf = make_summary_excel(problems_total, title="TOTAUX")
                zf.writestr("01_RAPPORT/extraits/PE_totaux_incorrects.xlsx", buf.getvalue())

        # 02_PDF -> Tous les PDFs renommés par PE (pas de sous-dossiers par erreur/statut)
        if include_pdfs:
            existing = set()
            idx_no_pe = 1

            for inv in invoices_data:
                pdf_bytes = inv.get("pdf_bytes")
                if not pdf_bytes:
                    continue

                pe = pe_of(inv)

                if pe:
                    zip_name = f"02_PDF/{safe(pe)}{pdf_suffix_am(inv)}.pdf"
                else:
                    zip_name = f"02_PDF/SANS_PE/facture_{idx_no_pe:03d}.pdf"
                    idx_no_pe += 1

                if zip_name in existing:
                    base = zip_name[:-4]
                    i = 2
                    while f"{base}_{i}.pdf" in existing:
                        i += 1
                    zip_name = f"{base}_{i}.pdf"
                existing.add(zip_name)

                zf.writestr(zip_name, pdf_bytes)

        # 03_ERREURS_PAR_PE
        if include_erreurs_par_pe:
            errors_rows = build_errors_rows()
            errors_by_pe = {}
            for r in errors_rows:
                errors_by_pe.setdefault(r["PE"], []).append(r)

            for inv in invoices_data:
                pe = pe_of(inv)
                if not pe:
                    continue

                has_any_issue = (
                    (inv.get("statut_final") or "").upper() != "OK"
                    or bool(inv.get("flags"))
                    or bool(inv.get("line_comparison"))
                )
                if not has_any_issue:
                    continue

                pe_clean = safe(pe)
                base = f"03_ERREURS_PAR_PE/{pe_clean}"

                pdf_bytes = inv.get("pdf_bytes")
                if pdf_bytes:
                    zf.writestr(f"{base}/{pe_clean}.pdf", pdf_bytes)

                try:
                    xbuf = make_total_qty_error_excel_for_invoice(inv)
                    zf.writestr(f"{base}/erreur_qte_total_{pe_clean}.xlsx", xbuf.getvalue())
                except:
                    pass

                if has_price_issue(inv):
                    try:
                        xbuf2 = make_price_error_excel_for_invoice(inv)
                        zf.writestr(f"{base}/ecart_prix_lignes_{pe_clean}.xlsx", xbuf2.getvalue())
                    except:
                        pass

                pe_rows = errors_by_pe.get(pe, [])
                if pe_rows:
                    df_pe = pd.DataFrame(pe_rows)
                    b = BytesIO()
                    with pd.ExcelWriter(b, engine="openpyxl") as w:
                        df_pe.to_excel(w, sheet_name="ERREURS", index=False)
                    b.seek(0)
                    zf.writestr(f"{base}/erreur_resume_{pe_clean}.xlsx", b.getvalue())

        # 04_CSV
        if include_csv:
            items_buf = export_items_to_csv(invoices_data, csv_folder_in_zip="04_CSV")
            zf.writestr("04_CSV/lignes_articles.csv", items_buf.getvalue())

            err_rows = build_errors_rows()
            df_err = pd.DataFrame(err_rows)
            b = BytesIO()
            df_err.to_csv(b, index=False, encoding="utf-8-sig", sep=";")
            b.seek(0)
            zf.writestr("04_CSV/erreurs.csv", b.getvalue())

    out.seek(0)
    return out




